"""マッピング表をXLSX形式で出力。

列構成（言及回数/総応答数の比率が低い順 → 検索ボリューム降順）:
1. プラットフォーム名
2. プロンプト
3. 検索ボリューム
4. 直近のファンアウトクエリ
5. ファンアウトクエリから読み取れる検索意図（AI解釈）
6. 言及（あり / 総応答）
7. 引用元URL（多い順）
8-13. site:検索 1〜3位（URL / タイトル）
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import common

COLUMNS = [
    ("プラットフォーム", 18),
    ("プロンプト", 40),
    ("検索ボリューム", 14),
    ("ファンアウトクエリ（直近）", 32),
    ("検索意図（AI解釈）", 40),
    ("言及（あり / 総応答）", 18),
    ("引用元URL（多い順）", 60),
    ("site:1位 URL", 32),
    ("site:1位 タイトル", 32),
    ("site:2位 URL", 32),
    ("site:2位 タイトル", 32),
    ("site:3位 URL", 32),
    ("site:3位 タイトル", 32),
]


def _set_header(ws) -> None:
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="E0E7EF", end_color="E0E7EF", fill_type="solid")
    for i, (name, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=i, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="left")


def _set_widths(ws) -> None:
    for i, (_, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _format_links(links: list) -> str:
    """引用元URL（多い順）を改行区切りで全件展開。

    各要素が dict なら url と count、str ならそのまま使う。
    count > 1 なら URL の後ろに `（×N）` を付ける。
    """
    lines = []
    for link in links:
        if isinstance(link, dict):
            url = link.get("url", "")
            if not url:
                continue
            count = link.get("count", 0)
            if count and count > 1:
                lines.append(f"{url}（×{count}）")
            else:
                lines.append(url)
        elif isinstance(link, str):
            lines.append(link)
    return "\n".join(lines)


def export_mapping(rows: list[dict], domain: str) -> Path:
    """rowsをXLSXファイルとして出力（既に build_mapping_rows でソート済みの前提）。

    各 row が持つべきキー:
    - platform: プラットフォーム名（chatgpt / perplexity）
    - prompt: プロンプト
    - fanout_query: 直近のファンアウトクエリ
    - intent: 検索意図（AI解釈）
    - mention_count: 自社が言及された応答数（int）
    - total_count: そのプロンプトの総応答数（int）
    - cited_pages: 引用元URL（多い順、全件）
    - site_search_results: site:検索結果（最大3件）
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "マッピング"

    _set_header(ws)
    _set_widths(ws)

    for r, row in enumerate(rows, start=2):
        site_results = row.get("site_search_results", [])
        site1 = site_results[0] if len(site_results) > 0 else {}
        site2 = site_results[1] if len(site_results) > 1 else {}
        site3 = site_results[2] if len(site_results) > 2 else {}

        mention_label = f"{row.get('mention_count', 0)} / {row.get('total_count', 0)}"
        values = [
            row.get("platform", ""),
            row.get("prompt", ""),
            row.get("volume", 0),
            row.get("fanout_query", ""),
            row.get("intent", ""),
            mention_label,
            _format_links(row.get("cited_pages", [])),
            site1.get("url", ""),
            site1.get("title", ""),
            site2.get("url", ""),
            site2.get("title", ""),
            site3.get("url", ""),
            site3.get("title", ""),
        ]
        for i, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = common.ensure_output_dir() / f"mapping_{domain}_{ts}.xlsx"
    wb.save(out_path)
    return out_path
