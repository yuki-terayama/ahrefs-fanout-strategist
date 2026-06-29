"""Phase 9: 深掘り分析と提言生成のための素材収集スクリプト。

Phase 8 で対策クエリ1件が選ばれたあと、以下を一括収集する：

1. Phase 4 と同じ target_info.md（検索意図解釈の素材）
2. cited_pages/: 引用元の本文を出現回数順に top_n 件
3. own_pages/: XLSX の site:1-3位 URL の自社既存ページ本文

すべて output/phase9/{slug}/ に集約。

Claude Code は出力ディレクトリ全体を読んで、Phase 9 の6ステップ
（解釈→引用元読了→自社ページ読了→既存改修/新規判断→提言生成→md保存）を対話で実施する。

使い方:
  python scripts/phase9_prepare.py \\
      --brand-name "Ahrefs" --brand-url "ahrefs.com" \\
      --prompt "無料のSEO分析ツールは？" \\
      --platform "perplexity" \\
      --xlsx output/mapping_ahrefs_com_20260628_203923.xlsx \\
      --top-cited 5
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook

from scripts.lib import aggregation, common, csv_loader, playwright_fetch
from scripts.main import DEFAULT_COMPETITOR_PHRASES
from scripts.phase4_intent import build_target_info_md, to_slug


def _safe_filename(url: str, max_len: int = 80) -> str:
    """URL をファイル名に使える形に整形"""
    s = re.sub(r"^https?://", "", url)
    s = re.sub(r'[\\/:*?"<>|\n\r\t]', "_", s)
    return s[:max_len]


def read_site_search_urls_from_xlsx(
    xlsx_path: Path, prompt: str, platform: str
) -> list[dict]:
    """XLSX から指定 (prompt, platform) の行を見つけて site:1-3位 URL を抽出"""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    header = [c.value for c in ws[1]]

    def col_idx(name: str) -> int:
        return header.index(name)

    idx_platform = col_idx("プラットフォーム")
    idx_prompt = col_idx("プロンプト")
    rank_cols = [
        (col_idx("site:1位 URL"), col_idx("site:1位 タイトル")),
        (col_idx("site:2位 URL"), col_idx("site:2位 タイトル")),
        (col_idx("site:3位 URL"), col_idx("site:3位 タイトル")),
    ]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx_platform] == platform and row[idx_prompt] == prompt:
            urls = []
            for url_i, title_i in rank_cols:
                url = row[url_i]
                title = row[title_i] or ""
                if url:
                    urls.append({"url": url, "title": title})
            return urls
    return []


def run(
    brand_name: str,
    brand_url: str,
    prompt: str,
    platform: str,
    xlsx_path: Path,
    top_cited: int,
    input_dir: Path,
) -> int:
    slug = to_slug(prompt, platform)
    out_dir = common.ensure_output_dir(f"phase9/{slug}")
    cited_dir = out_dir / "cited_pages"
    own_dir = out_dir / "own_pages"
    cited_dir.mkdir(exist_ok=True)
    own_dir.mkdir(exist_ok=True)

    print(f"=== Phase 9: 素材収集 ===")
    print(f"対象: {platform} / {prompt}")
    print(f"出力先: {out_dir}")

    # 1. CSV ロード + スクリーニング
    print(f"\n[1/4] CSV ロード ({input_dir})")
    all_responses = csv_loader.load_brand_radar_csvs(input_dir)
    print(f"  総応答数（重複除去後）: {len(all_responses)}")
    screened = [
        r
        for r in all_responses
        if not any(c in (r.get("question") or "").lower() for c in DEFAULT_COMPETITOR_PHRASES)
    ]
    print(f"  競合名スクリーニング後: {len(screened)}")

    # 2. 対象応答抽出
    target_resps = [
        r
        for r in screened
        if r.get("question") == prompt and r.get("data_source") == platform
    ]
    print(f"\n[2/4] 対象応答抽出: {len(target_resps)} 件")
    if not target_resps:
        print(f"  ERROR: 対象応答なし")
        return 1

    # 3. target_info.md 生成（Phase 4 と同じロジックを再利用）
    print(f"\n[3/4] target_info.md 生成（検索意図解釈用素材）")
    md = build_target_info_md(prompt, platform, brand_name, target_resps)
    (out_dir / "target_info.md").write_text(md, encoding="utf-8")
    print(f"  保存: {(out_dir / 'target_info.md').name}")

    # 4. 引用元 top_n + 自社既存ページ本文を一括取得
    print(f"\n[4/4] 引用元トップ{top_cited} + 自社既存ページ 本文取得")

    cited = aggregation.cited_pages_by_query(target_resps, top_n=top_cited)
    cited_list = cited.get((prompt, platform), [])
    print(f"  引用元 (cited, top {top_cited}): {len(cited_list)} 件")

    # Fallback: cited が 0件 のとき Found but not cited を代用
    referenced_list: list[dict] = []
    if cited_list:
        referenced_list = [{**c, "kind": "cited"} for c in cited_list]
    else:
        from collections import Counter
        fnc_counter: Counter = Counter()
        for r in target_resps:
            for li in r.get("found_not_cited") or []:
                url = li.get("url") if isinstance(li, dict) else None
                if url:
                    fnc_counter[url] += 1
        fnc_top = fnc_counter.most_common(top_cited)
        referenced_list = [
            {"url": u, "title": "", "count": c, "kind": "found_not_cited"}
            for u, c in fnc_top
        ]
        print(
            f"  ⚠ cited が0件のため Found but not cited を fallback として使用: {len(referenced_list)} 件"
        )

    for i, c in enumerate(referenced_list, 1):
        print(f"    [{i}] x{c['count']} ({c['kind']})  {c['url']}")

    own_urls = read_site_search_urls_from_xlsx(xlsx_path, prompt, platform)
    print(f"  自社既存ページ (site: 1-3位): {len(own_urls)} 件")
    for i, o in enumerate(own_urls, 1):
        print(f"    [{i}] {o['url']}")

    all_urls = [r["url"] for r in referenced_list] + [o["url"] for o in own_urls]
    if not all_urls:
        print(f"  取得対象URLなし。スキップ")
    else:
        print(f"\n  Playwright で本文取得中（{len(all_urls)} URL、1セッション）...")
        pages = playwright_fetch.fetch_pages_text(all_urls)

        # 引用元保存（cited / found_not_cited を kind 列で識別）
        for i, c in enumerate(referenced_list, 1):
            url = c["url"]
            kind = c["kind"]
            text = pages.get(url, "")
            fname = f"{i:02d}_{kind}_x{c['count']}_{_safe_filename(url)}.txt"
            content = (
                f"URL: {url}\n"
                f"種別: {kind}\n"
                f"出現回数: {c['count']}\n"
                f"取得文字数: {len(text)}\n"
                f"---\n\n"
                f"{text[:30_000]}"
            )
            (cited_dir / fname).write_text(content, encoding="utf-8")
            print(f"    referenced [{i}] {kind:<16} {len(text):>6} chars → {fname}")

        # 自社既存ページ保存
        for i, o in enumerate(own_urls, 1):
            url = o["url"]
            title = o["title"]
            text = pages.get(url, "")
            fname = f"{i:02d}_{_safe_filename(url)}.txt"
            content = (
                f"URL: {url}\n"
                f"タイトル: {title}\n"
                f"取得文字数: {len(text)}\n"
                f"---\n\n"
                f"{text[:30_000]}"
            )
            (own_dir / fname).write_text(content, encoding="utf-8")
            print(f"    own        [{i}] {len(text):>6} chars → {fname}")

    print(f"\n=== 完了 ===")
    print(f"次のアクション: Claude Code が以下を読んで Phase 9 の判断・提言を生成")
    print(f"  - {out_dir.relative_to(common.PROJECT_ROOT)}/target_info.md  （検索意図解釈）")
    print(f"  - {out_dir.relative_to(common.PROJECT_ROOT)}/cited_pages/    （引用元本文）")
    print(f"  - {out_dir.relative_to(common.PROJECT_ROOT)}/own_pages/      （自社既存ページ本文）")
    return 0


def main() -> int:
    p = argparse.ArgumentParser(description="Phase 9 素材収集")
    p.add_argument("--brand-name", required=True)
    p.add_argument("--brand-url", required=True)
    p.add_argument("--prompt", required=True)
    p.add_argument("--platform", required=True, choices=["chatgpt", "perplexity"])
    p.add_argument("--xlsx", required=True, help="Phase 7 で出力した mapping XLSX")
    p.add_argument("--top-cited", type=int, default=5)
    p.add_argument("--input-dir", default="input")
    args = p.parse_args()
    return run(
        brand_name=args.brand_name,
        brand_url=args.brand_url,
        prompt=args.prompt,
        platform=args.platform,
        xlsx_path=Path(args.xlsx),
        top_cited=args.top_cited,
        input_dir=Path(args.input_dir),
    )


if __name__ == "__main__":
    sys.exit(main())
